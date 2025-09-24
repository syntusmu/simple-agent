#!/usr/bin/env python3
"""
Enhanced Excel Loader for processing Excel files with advanced capabilities.

This module provides comprehensive Excel loading functionality:
- Identifies and extracts single-row information (titles, notes, etc.)
- Handles merged cells and multi-row headers properly
- Avoids header duplication in merged cells
- Outputs data in header:row format with one row per chunk
- Separates text content from table data

Classes:
    ExcelLoader: Enhanced loader class for Excel files with advanced processing

Functions:
    load_excel(): Convenience function for loading Excel files
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Union, Tuple

import openpyxl
import pandas as pd
from langchain.schema import Document
from openpyxl.worksheet.worksheet import Worksheet

# Configure logger
logger = logging.getLogger(__name__)


class ExcelLoader:
    """
    Enhanced Excel loader that handles single-row information and table data separately.
    
    This loader provides advanced functionality for processing Excel files:
    - Identifies single-row information (titles, notes, etc.)
    - Extracts text content separately from table data
    - Handles merged cells without duplication
    - Outputs data in header:row format
    - One row per chunk for better document splitting
    
    Attributes:
        file_path: Path to the Excel file
        sheet_name: Name or index of the sheet to process (None for active sheet)
        preserve_merged_cells: Whether to preserve merged cell content
        header_rows: List of 0-based row indices for multi-row headers
        data_start_row: 0-based row index where data starts
        header_row: Single header row (1-indexed) for simple cases
        backend: Processing backend ('pandas' or 'openpyxl')
    """
    
    def __init__(
        self, 
        file_path: str, 
        sheet_name: Optional[str] = None,
        preserve_merged_cells: bool = True,
        data_start_row: Optional[int] = None,
        header_row: Optional[int] = None,
        header_rows: Optional[List[int]] = None,
        backend: str = 'auto'
    ):
        """
        Initialize the enhanced Excel loader.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Sheet name or index (None for active sheet)
            preserve_merged_cells: Whether to preserve merged cell content
            header_rows: List of 0-based row indices for multi-row headers
            data_start_row: 0-based row index where data starts
            header_row: Single header row (1-indexed) for simple cases
            backend: Processing backend ('pandas', 'openpyxl', or 'auto')
        """
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name
        self.preserve_merged_cells = preserve_merged_cells
        self.header_rows = header_rows
        self.data_start_row = data_start_row
        self.header_row = header_row
        self.backend = backend
        
        self._validate_file()
        self._configure_loader()

    # Validation methods
    def _validate_file(self) -> None:
        """Validate that the Excel file exists and is accessible."""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

    # Configuration methods
    def _configure_loader(self) -> None:
        """Configure the loader based on provided parameters."""
        # Use openpyxl for better merged cell handling
        self.backend = 'openpyxl'
        
        # Configure processing mode based on parameters
        if self.header_rows is None and self.header_row is None:
            # Auto-detect: assume rows 3-5 are headers (0-based: 2-4)
            self.header_rows = [2, 3, 4]
            self.data_start_row = 5
            self.mode = 'enhanced'
        elif self.header_rows is not None:
            self.mode = 'multi_header'
            if self.data_start_row is None:
                self.data_start_row = max(self.header_rows) + 1
        elif self.header_row is not None:
            self.mode = 'single_header'
            self.header_rows = [self.header_row - 1]  # Convert to 0-based
            if self.data_start_row is None:
                self.data_start_row = self.header_row
        else:
            self.mode = 'enhanced'
            self.header_rows = [2, 3, 4]
            self.data_start_row = 5

    # Main loading methods
    def load(self) -> List[Document]:
        """
        Load documents from Excel file with enhanced processing.
        
        Returns:
            List of Document objects including text content and table data
        """
        logger.info(f"Loading Excel file: {self.file_path} (mode: {self.mode}, backend: {self.backend})")
        return self._load_with_enhanced_processing()

    def _load_with_enhanced_processing(self) -> List[Document]:
        """Load using enhanced processing with single-row detection."""
        try:
            # Load workbook and select worksheet
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws, actual_sheet_name = self._select_worksheet(wb)
            
            logger.info(f"Processing sheet: {actual_sheet_name}")
            
            # Process merged cells if needed
            if self.preserve_merged_cells:
                self._fill_merged_cells(ws)
            
            # Extract content components
            single_row_info = self._extract_single_row_information(ws)
            headers = self._extract_headers(ws)
            data_rows = self._extract_data_rows(ws, headers)
            
            # Create documents
            documents = self._create_documents(single_row_info, data_rows, actual_sheet_name)
            
            wb.close()
            logger.info(f"Successfully loaded {len(documents)} documents")
            return documents
            
        except Exception as e:
            logger.error(f"Error loading Excel with enhanced processing: {e}")
            raise

    def _select_worksheet(self, wb: openpyxl.Workbook) -> Tuple[Worksheet, str]:
        """Select and return the appropriate worksheet."""
        if self.sheet_name:
            if isinstance(self.sheet_name, str):
                ws = wb[self.sheet_name]
                actual_sheet_name = self.sheet_name
            else:
                sheet_names = wb.sheetnames
                ws = wb[sheet_names[self.sheet_name]]
                actual_sheet_name = sheet_names[self.sheet_name]
        else:
            ws = wb.active
            actual_sheet_name = ws.title
        
        return ws, actual_sheet_name

    # Data extraction methods
    def _extract_single_row_information(self, ws: Worksheet) -> List[str]:
        """
        Identify rows that contain only single-cell information (titles, notes, etc.).
        
        Args:
            ws: Worksheet object
            
        Returns:
            List of single-row information strings
        """
        single_row_info = []
        max_header_row = max(self.header_rows) if self.header_rows else 0
        
        for row_idx in range(1, max_header_row + 1):  # 1-based indexing
            non_empty_cells = 0
            non_empty_content = []
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                if cell_value is not None and str(cell_value).strip():
                    non_empty_cells += 1
                    content = str(cell_value).strip()
                    if content not in non_empty_content:  # Avoid duplicates
                        non_empty_content.append(content)
            
            # Determine if this is single-row information
            is_single_info = self._is_single_row_info(non_empty_cells, non_empty_content, ws.max_column)
            
            if is_single_info and non_empty_content:
                info_text = " ".join(non_empty_content)
                single_row_info.append(info_text)
                logger.info(f"Found single-row info at row {row_idx}: {info_text}")
        
        return single_row_info

    def _is_single_row_info(self, non_empty_cells: int, non_empty_content: List[str], max_column: int) -> bool:
        """Determine if a row contains single-row information."""
        return (
            non_empty_cells == 1 or 
            (non_empty_cells > 0 and len(set(non_empty_content)) == 1) or
            (non_empty_cells > 0 and non_empty_cells <= 3 and max_column > 6)
        )

    def _extract_headers(self, ws: Worksheet) -> List[str]:
        """
        Extract headers from header rows, avoiding duplication from merged cells.
        
        Args:
            ws: Worksheet object
            
        Returns:
            List of combined header strings
        """
        if not self.header_rows:
            return []
        
        headers = []
        
        for col_idx in range(1, ws.max_column + 1):
            header_parts = []
            
            for row_idx in self.header_rows:
                cell = ws.cell(row=row_idx + 1, column=col_idx)  # Convert to 1-based
                cell_value = cell.value
                if cell_value is not None:
                    content = str(cell_value).strip().replace('\n', ' ')  # Remove newlines
                    if content and content not in header_parts:  # Avoid duplicates
                        header_parts.append(content)
            
            # Combine header parts or create default
            if header_parts:
                combined_header = " | ".join(header_parts)
                headers.append(combined_header)
            else:
                headers.append(f"Column_{col_idx}")
        
        logger.info(f"Extracted {len(headers)} headers")
        return headers

    def _extract_data_rows(self, ws: Worksheet, headers: List[str]) -> List[Dict[str, str]]:
        """
        Extract data rows in enhanced format.
        
        Args:
            ws: Worksheet object
            headers: List of header strings
            
        Returns:
            List of dictionaries with header:value pairs
        """
        data_rows = []
        
        for row_idx in range(self.data_start_row + 1, ws.max_row + 1):  # Convert to 1-based
            row_data = {}
            has_data = False
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                
                if cell_value is not None:
                    value = str(cell_value).strip()
                    if value:
                        row_data[header] = value
                        has_data = True
                    else:
                        row_data[header] = ""
                else:
                    row_data[header] = ""
            
            if has_data:
                data_rows.append(row_data)
        
        logger.info(f"Extracted {len(data_rows)} data rows")
        return data_rows

    # Document creation methods
    def _create_documents(self, single_row_info: List[str], data_rows: List[Dict[str, str]], sheet_name: str) -> List[Document]:
        """Create Document objects from extracted data."""
        documents = []
        
        # Add text content document if single-row info exists
        if single_row_info:
            text_content = "\n".join(single_row_info)
            text_doc = Document(
                page_content=text_content,
                metadata={
                    "source": str(self.file_path),
                    "sheet_name": sheet_name,
                    "content_type": "text_content",
                    "row_count": len(single_row_info)
                }
            )
            documents.append(text_doc)
        
        # Add table data documents (one per row)
        for i, row_data in enumerate(data_rows):
            content_parts = []
            for header, value in row_data.items():
                if value and str(value).strip():
                    content_parts.append(f'"{header}": "{value}"')
            
            if content_parts:
                content = "\n".join(content_parts)
                doc = Document(
                    page_content=content,
                    metadata={
                        "source": str(self.file_path),
                        "sheet_name": sheet_name,
                        "content_type": "table_data",
                        "row_index": i + self.data_start_row + 1,  # 1-based row number
                        "headers": list(row_data.keys())
                    }
                )
                documents.append(doc)
        
        return documents

    # Utility methods
    def _fill_merged_cells(self, ws: Worksheet) -> None:
        """
        Fill merged cells with the value from the top-left cell.
        
        Args:
            ws: Worksheet object
        """
        from openpyxl.cell.cell import MergedCell
        
        for merged_range in ws.merged_cells.ranges:
            # Get the value from the top-left cell
            top_left_cell = ws.cell(
                row=merged_range.min_row, 
                column=merged_range.min_col
            )
            merge_value = top_left_cell.value
            
            # Fill all cells in the merged range, but skip MergedCell objects
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    # Only set value for non-merged cells or create new cells
                    if not isinstance(cell, MergedCell):
                        cell.value = merge_value


def load_excel(
    file_path: str,
    sheet_name: Optional[str] = None,
    preserve_merged_cells: bool = True,
    header_rows: Optional[List[int]] = None,
    data_start_row: Optional[int] = None,
    header_row: Optional[int] = None,
    backend: str = 'auto'
) -> List[Document]:
    """
    Convenience function to load Excel file with enhanced processing.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Sheet name or index (None for active sheet)
        preserve_merged_cells: Whether to preserve merged cell content
        header_rows: List of 0-based row indices for multi-row headers
        data_start_row: 0-based row index where data starts
        header_row: Single header row (1-indexed) for simple cases
        backend: Processing backend ('pandas', 'openpyxl', or 'auto')
        
    Returns:
        List of Document objects
        
    Examples:
        # Simple single header
        documents = load_excel("data.xlsx", header_row=1)
        
        # Multi-row headers
        documents = load_excel("data.xlsx", header_rows=[1, 2, 3], data_start_row=4)
        
        # With merged cells
        documents = load_excel("data.xlsx", preserve_merged_cells=True)
    """
    loader = ExcelLoader(
        file_path=file_path,
        sheet_name=sheet_name,
        preserve_merged_cells=preserve_merged_cells,
        data_start_row=data_start_row,
        header_row=header_row,
        header_rows=header_rows,
        backend=backend
    )
    return loader.load()


# Backward compatibility aliases
ExcelMergedCellLoader = ExcelLoader
ExcelImprovedLoader = ExcelLoader
load_excel_with_merged_cells = load_excel
load_excel_improved = load_excel